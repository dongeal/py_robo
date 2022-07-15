from msilib.schema import Font
from numpy import size
import pyautogui
import pyperclip
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

underline = Border(bottom=Side(border_style='thin'))

def my_write(text):
    pyperclip.copy(text)
    pyautogui.hotkey("ctrl","v")

# 텍스트를 엘셀로 변환하여 work 에 sale_report.xlsx로 저장
pyautogui.hotkey("win","r")
pyautogui.sleep(0.2)
my_write(r"excel C:/work/판매금액.txt")
pyautogui.hotkey("enter")
pyautogui.sleep(5)
pyautogui.hotkey("F12")
pyautogui.sleep(1)
pyautogui.write("sale_report")
pyautogui.sleep(0.5)
pyautogui.hotkey("alt","t")
pyautogui.sleep(0.2)
pyautogui.hotkey('down')
pyautogui.sleep(0.2)
pyautogui.hotkey('pageup')
pyautogui.sleep(0.2)
pyautogui.hotkey("enter")
pyautogui.sleep(0.2)
pyautogui.hotkey("enter")
pyautogui.sleep(0.2)
pyautogui.hotkey("y")
pyautogui.sleep(0.2)
pyautogui.hotkey("alt","F4")

pyautogui.sleep(2)

# 엑셀파일 불러오기
from openpyxl import load_workbook
wb = load_workbook(r"c:/work/sale_report.xlsx")
ws = wb.active # 활성화된 Sheet
ws.insert_cols(2,1)


maxrow=ws.max_row
ws.move_range(f'L1:L{maxrow}',rows=0, cols=-10)


# # 총계열 구하기
r=ws.max_row
row_from =5
row_total = 0
for cell in ws['A']:
    row_total += 1
    a=str(cell.value)
    if a[0]=='총':
            break


# sorting 할 셀 데이타 읽기
row_to =row_total-2   # 총계열 구한것에서 -2
row_range=ws[row_from:row_to]
rdata=[]
tdata=[]
for row in row_range:
    for cell in row:
        rdata.append(cell.value) # 행 data [달성, 1000,500,3000, ...]
 
    tdata.append(rdata) # [[rdata1],[rdata2]....] 행 들로 이루어진 2차원 배열
    rdata=[] # 다시 다음 행으로 가기전 행 data 비우기

sdata =sorted(tdata,key=lambda x:x[1],reverse=True) # sorting 기준열 (2)-1=1

x_underline=0 # 5 줄마다 언더라인 그릴 준비

for x  in range(5,row_to+1): #  (5 , 총계열 구한것 -1)
    x_underline += 1
    for y in range(1,12):    
        ws.cell(row=x,column=y,value=sdata[x-5][y-1]) #cell에 sorting 된 data 쓰기
        ws.cell(row=x,column=y).number_format='#,##0'
        if x_underline%5==0:
            ws.cell(row=x,column=y).border=underline
# 1열에 숫자 넣기
ws.insert_cols(1,1)
n = 0
for x in range(5,row_to+1):
    n += 1
    ws.cell (row=x,column=1, value=n)

    
# 모양 다듬기
for x in range(1,100):
    ws.row_dimensions[x].height = 12
    for y in range(1,100):
        ws.cell(x,y).font =Font(name='굴림체', size=10)

title= ws['E1'].value +ws['F1'].value+ws['G1'].value+ws['H1'].value
ws.merge_cells(start_row=1,start_column=5,end_row=1,end_column=10)
ws['E1'].value = title
ws['E1'].border= underline
ws['F1'].border= underline
ws['G1'].border= underline
ws['H1'].border= underline
ws['I1'].border= underline

ws['E1'].font=Font(name='굴림체',size=12,bold=True)

title2=ws.cell(row=row_total+2,column=5).value+ws.cell(row=row_total+2,column=6).value
t2=row_total+2
ws.merge_cells(start_row=t2,start_column=5,end_row=t2,end_column=7)
ws.cell(row=t2,column=5).value = title2
ws.cell(row=t2,column=5).border=underline
ws.cell(row=t2,column=6).border=underline
ws.cell(row=t2,column=7).border=underline
ws.cell(row=t2,column=5).font=Font(name='굴림체',size=12,bold=True)

ws.column_dimensions['A'].width =3
ws.column_dimensions['B'].width =14
ws.column_dimensions['C'].width =13
ws.column_dimensions['D'].width =11
ws.column_dimensions['E'].width =11
ws.column_dimensions['F'].width =11
ws.column_dimensions['G'].width =11
ws.column_dimensions['H'].width =11
ws.column_dimensions['I'].width =11
ws.column_dimensions['J'].width =11
ws.column_dimensions['K'].width =11
ws.column_dimensions['L'].width =11

test = ws.cell(row=row_total+4,column=2).value
if test[0]=='-': # 수출이 없는경우  
    ws.delete_rows(row_total+2,7)

wb.save(r"c:/work/sale_report.xlsx")


