import pyautogui
import pyperclip
from openpyxl import load_workbook

wb = load_workbook(r"sale_report.xlsx")
ws = wb.active # 활성화된 Sheet

row_from=5
row_to =36   # 총계열 구한것에서 -2
row_range=ws[row_from:row_to]
rdata=[]
tdata=[]
for row in row_range:
    for cell in row:
        rdata.append(cell.value) 
 
    tdata.append(rdata)
    rdata=[]
# print(tdata)
sdata =sorted(tdata,key=lambda x:x[10],reverse=True) # 쏘팅열 -1
# print(" ")
# print(sdata)

for x  in range(5,37): #  (5 , 총계열 구한것 -1)
    for y in range(1,12):    
        ws.cell(row=x,column=y,value=sdata[x-5][y-1]) #


wb.save("sale_sorted.xlsx")

