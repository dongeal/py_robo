from openpyxl import load_workbook

wb=load_workbook('sale_report.xlsx')
ws=wb.active
mr=ws.max_row
row_from =5
row_to = 0
for cell in ws['A']:
    row_to += 1
    a=str(cell.value)
    if a[0]=='총':
            break


print(row_to)

ws.move_range(f'k1:k{mr}',rows=0, cols=-9)




wb.save('sale_modified.xlsx')