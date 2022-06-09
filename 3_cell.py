from openpyxl import Workbook
wb= Workbook()
ws = wb.active
ws.title ="NadoSheet"

ws["A1"]=1
ws["A2"]=2
ws["A3"]=3


ws["B1"]=4
ws["B2"]=5
ws["B3"]=6

print(ws["A1"])
print(ws["A1"].value)
print(ws["A10"].value) # 없으면 None

print(ws.cell(column=1,row=1).value)
print(ws.cell(column=2,row=1).value)

c=ws.cell(column=3,row=1,value=10)
print(c.value)

from random import *
index =1
for x in range(1,11): # 10개의 row
    for y in range(1, 11): # 10 개의 column
        # ws.cell(row=x,column=y, value= randint(0,100)) # 0-100
        ws.cell(row=x,column=y,value=index)
        index +=1

wb.save("sample.xlsx")