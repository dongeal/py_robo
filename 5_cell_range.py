from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

ws.append(["번호","영어","수학"])
for i in range(1,11):
    ws.append([i,randint(0,100),randint(0,100)])

# col_B = ws["B"]
# # print(col_B)

# for cl in col_B:
#     print(cl.value)


# col_range =ws["B:C"]
# for cols in col_range:
#     for cl in cols:
#         print(cl.value)
# row_title = ws[1]
# for cl in row_title:
#     print(cl.value, end=" ")
# print()

# row_range = ws[2:6]
# for rows in row_range:
#     for cl in rows:
#         print (cl.value, end="    ")
#     print()

# row_range =ws[2:ws.max_row]
# for rows in row_range:
#     for cl in rows:
#         print (cl.value, end=" ")
#     print()
from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row]
# for rows in row_range:
#     for cl in rows:
#         # print(cl.value, end=" ")
#         # print(cl.coordinate, end=" ")
#         xy= coordinate_from_string(cl.coordinate)
#         #print(xy, end=" ")
#         print(xy[0],end="/")
#         print(xy[1], end=" ")
#     print()
# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[2].value)

#print(tuple(ws.columns))
# for col in tuple(ws.columns):
#     print(col[0].value)


# for row in ws.iter_rows(): # 전체 로우
#     print(row[0].value)
for col in ws.iter_cols(): # 전체 컬럼
    print(col[0].value)

# for row in ws.iter_rows(min_row=2,max_row=11,min_col=2,max_col=3): 
#     print(row[0].value,row[1].value)
for col in ws.iter_cols(min_row=1,max_row=11,min_col=1,max_col=3): 
    print(col[0].value,col[1].value,col[2].value,col[3].value)


wb.save("sample.xlsx")