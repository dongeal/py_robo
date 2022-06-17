from openpyxl import Workbook
wb=Workbook()
ws= wb.active

ws.append(("학번","출석","퀴즈1","퀴즈2","중간고사","기말고사","프로젝트"))

scores=[
(1,	10,	8,	5,	14,	26,	12),
(2,	7,	3,	7,	15,	24,	18),
(3,	9,	5,	8,	8, 12,	4),
(4,	7,	8,	7,	17,	21,	18),
(5,	7,	8,	7,	16,	25,	15),
(6,	3,	5,	8,	8,	17,	0),
(7,	4,	9,	10,	16,	27,	18),
(8,	6,	6,	6,	15,	19,	17),
(9,	10,	10,	9,	19,	30,	19),
(10, 9,	8,	8,	20,	25,	20)]

for s in scores:
    ws.append(s)

# 1.퀴즈2를 전부 10점으로
for row in range(2, ws.max_row+1):
    ws["D"+str(row)].value = 10

# 2. H열에 총점,I 열에 성적 삽입
ws["H1"]="총점"
ws["I1"]="성적"
for r in range(2, ws.max_row+1):
    ws.cell(row=r,column=8).value = "=SUM(B{}:G{})".format(r,r)

wb.save("scores.xlsx")   

import win32com.client
# pip install pywin32

def LoadExcelFunction():
    #win32com
    #excel 사용할 수 있게 설정
    excel = win32com.client.Dispatch("Excel.Application")
    #임시 Workbook 객체 생성 및 엑셀 열기
    temp_wb = excel.Workbooks.Open(
        r"C:\Users\ilheu\Desktop\py_rpa\1_exel\scores.xlsx")
    #저장
    temp_wb.Save()
    #excel 종료
    excel.quit()
    
LoadExcelFunction()

from openpyxl import load_workbook
wb=load_workbook("scores.xlsx",data_only=True)
ws=wb.active     

for r in range(2, ws.max_row+1):
    if ws.cell(row=r,column=2).value < 5 :
        ws.cell(row=r,column=9).value ="F"
        continue
    if ws.cell(row=r,column=8).value >= 90:
        ws.cell(row=r,column=9).value ="A"
        continue
    if ws.cell(row=r,column=8).value >= 80:
        ws.cell(row=r,column=9).value ="B"
        continue
    if ws.cell(row=r,column=8).value >= 70:
        ws.cell(row=r,column=9).value ="C"
        continue
    if ws.cell(row=r,column=8).value >= 60:
        ws.cell(row=r,column=9).value ="D"
        continue
    else:
        ws.cell(row=r,column=9).value ="D"

wb.save("scores.xlsx")   



