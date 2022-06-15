from openpyxl import load_workbook

# 수식그대로 가져오기
# wb=load_workbook("sample_formula.xlsx")
# ws=wb.active

# for row in ws.values:
#     for cell in row:
#         print(cell)


# 데이터상태로가져오기
# evaluate되지 않은 데이터는 None로 표시됨
wb=load_workbook("sample_formula.xlsx", data_only=True)
ws=wb.active

for row in ws.values:
    for cell in row:
        print(cell)
