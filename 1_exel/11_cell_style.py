from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment,PatternFill
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호, 영어, 수학
a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

# 넓이 와 높이 설정
ws.column_dimensions["A"].width = 5
ws.row_dimensions[1].height = 50

#폰트
a1.font=Font(color="FF0000", italic=True, bold= True)
b1.font=Font(color="00ff00",name="Arial", strike= True)
c1.font=Font(color="0000ff",underline="double",size=20)

#테두리
a1.border=Border(left=Side(style="thick"),right=Side(style="thin"),
        top=Side(style="double"),bottom=Side("thin"))

custom_border=Border(left=Side(style="thick"),right=Side(style="thin"),
        top=Side(style="double"),bottom=Side("thin"))
b1.border= custom_border
c1.border= custom_border

#90점 넘으면 초록색
for row in ws.rows:        
        for cell in row:
                cell.alignment=Alignment(horizontal="center",vertical="center")

                if cell.column ==1 : # "번호열 제외"
                        continue

                #cell이 정수 90 점보다큼
                if isinstance(cell.value, int) and cell.value > 90:
                        cell.fill= PatternFill(fgColor="00ff00",fill_type="solid")
                        cell.font=Font(color="ff0000")


# 틀 고정
ws.freeze_panes="B2"


wb.save("sample_style.xlsx")