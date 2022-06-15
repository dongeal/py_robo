import openpyxl


from openpyxl import load_workbook
wb = load_workbook("sample_merge.xlsx")
ws = wb.active

# B2:D2  병합셀 해지
ws.unmerge_cells("B2:D2")

wb.save("sample_unmerge.xlsx")